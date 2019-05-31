

from openpyxl import *
from deap import base
from deap import creator
from deap import tools
import collections

class Classes:
    def __init__(self, Name):
        self.ClassName = Name
    def __eq__(self, other):
        if not isinstance(other, Classes):
            return 0
        return self.ClassName == other.ClassName

class Term:
    def __init__(self, className,  day, time):
        self.Day = day
        self.TimeBlock = time
        self.ClassName = className
    def __eq__(self, other):
        if not isinstance(other, Term):
            return 0
        return self.Day==other.Day and self.TimeBlock==other.TimeBlock and self.ClassName==other.ClassName
    def __str__(self):
        switchdays = {
            1:"Monday",
            2:"Tuesday",
            3:"Wednesday",
            4:"Thursday",
            5:"Friday",
        }
        return self.ClassName + ", " + switchdays.get(self.Day) + ", block: " + str(self.TimeBlock)



class Person:
    FirstName = None
    LastName = None
    Preferences = []
    ClassTerms = []
    def __init__(self, firstName, lastName):
        self.FirstName=firstName
        self.LastName=lastName
        self.Preferences = []
        self.ClassTerms = []
    def __str__(self):
        return self.FirstName + " " + self.LastName

def Main():
    peopleWorkbook = load_workbook(filename = 'people.xlsx')
    people = peopleWorkbook.active
    classesWorkbook = load_workbook(filename = 'classes.xlsx')
    classes = classesWorkbook.active
    terms = []
    persons = []
    for row in classes.iter_rows(min_row=2):
        terms.append(Term(row[1].value, row[2].value, row[3].value))
    for row in people.iter_rows(min_row=2):
        person = Person(row[1].value, row[2].value)
        for x in range(5):
            for y in range(6):
                person.Preferences.append(((x, y), row[3+(6*x)+y].value))
        persons.append(person)
    creator.create("FitnessMin", base.Fitness, weights=(1.0,))
    creator.create("Individual", list, fitness=creator.FitnessMin)





Main()
if __name__=="__Main__":
    Main()

